import openpyxl

path="Path of your ecxel sheet in local machine"

workbook=openpyxl.load_workbook(path)


# when you only have one sheet in excel                 when you have multiple sheets in excel
sheet=workbook.active                                   #workbook.get_sheet_by_name("Sheet1")


rows=sheet.max_row
cols=sheet.max_column

print(rows)
print(cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(row=r,column=c).value,end="     ")

    print()
